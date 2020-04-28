import xlsxwriter
import xlrd
import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_NONE, BORDER_MEDIUM, DEFAULT_BORDER
import json
from os import walk

abc = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


reg_keys = {
    'msk': 'Москва',
    'nsk': 'Новосибирск',
    'spb': 'С-Петербург',
    'ekb': 'Екатеринбург',
    'vn': 'Великий Новгород',
    'barnaul': 'Барнаул',
    'arkhangelsk': 'Архагельск',
    'astrakhan': 'Астрахань',
    'chelyabinsk': 'Челябинск',
    'kaliningrad': 'Калининград',
    'vladivostok': 'Владивосток',
    'biysk': 'Бийск',
    'blag': 'Благовещенск',
    'bratsk': 'Братск',
    'bryansk': 'Брянск',
    'vladimir': 'Владимир',
    'volgograd': 'Волгоград',
    'vologda': 'Вологда',
    'voronezh': 'Воронеж',
    'gorno-altaisk': 'Горно-Алтайск',
    'ivanovo': 'Иваново',
    'izhevsk': 'Ижевск',
    'irkutsk': 'Иркутск',
    'kazan': 'Казань',
    'kaluga': 'Калуга',
    'kirov': 'Калуга',
    'komsomolsk': 'Комсомольск на Амуре',
    'kostroma': 'Кострома',
    'krasnodar': 'Краснодар',
    'kurgan': 'Курган',
    'kursk': 'Курск',
    'lipetsk': 'Липецк',
    'magnitogorsk': 'Магнитогорск',
    'makhachkala': 'Махачкала',
    'murmansk': 'Мурманск',
    'chelny': 'Набережные челны',
    'nizhnevartovsk': 'Нижневартовск',
    'nn': 'Нижний новгород',
    'tagil': 'Тагил',
    'novokuznetsk': 'Новокузнецк',
    'novorossiysk': 'Новороссийск',
    'omsk': 'Омск',
    'orenburg': 'Оренбург',
    'orel': 'Орел',
    'penza': 'Пенза',
    'perm': 'Пермь',
    'pskov': 'Псков',
    'rostov': 'Ростов на дону',
    'ryazan': 'Рязань',
    'samara': 'Самара',
    'saransk': 'Саранск',
    'saratov': 'Саратов',
    'smolensk': 'Смоленск',
    'sochi': 'Сочи',
    'surgut': 'Сургут',
    'syktyvkar': 'syktyvkar',
    'tambov': 'Тамбов',
    'tver': 'Тверь',
    'togliatti': 'Тольятти',
    'tomsk': 'Томск',
    'tula': 'Тула',
    'tyumen': 'Тюмень',
    'ulan-ude': 'Улан Уде',
    'ulyanovsk': 'Ульяновск',
    'ufa': 'Уфа',
    'khabarovsk': 'Хабаровск',
    'cheboksary': 'Чебоксары',
    'chita': 'Чита',
    'yuzhnosakhalinsk': 'yuzhnosakhalinsk',
    'yakutsk': 'Якутск',
    'yaroslavl': 'Ярославль'
}

def set_cell(col, num):
    return '{0}{1}'.format(abc[col], num)


def write_cell(cell, value, sheet):
    try:
        sheet[cell] = value
    except Exception:
        sheet[cell] = str(value)


def write_arr_cell(colume, line, arr, sheet):
    n = line
    for i in arr:
        write_cell(set_cell(colume, n), i, sheet)
        n += 1


def create_file_from_dict_list(name, data_all):
    file_name = 'exele_files/{0}.xlsx'.format(name)
    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet(name='Page1')
    workbook.close()
    headers = ['Город', 'Название', 'Рейтинг', 'Телефон', 'Адрес', 'Время работы', 'Ценовая категория', 'Сотрудники', 'Сайт', 'Акции', 'Теги']
    xfile = openpyxl.load_workbook(file_name)
    sheet = xfile.get_sheet_by_name('Page1')
    line = 1
    col = 1
    for data in data_all:
        m_var = 0
        for key in data:
            if type(data[key]) == type([5, 8]):
                if len(data[key]) > m_var:
                    m_var = len(data[key])
                try:
                    if type(data[key][0]) != type({'d': 8, 'l': 9}):
                        write_arr_cell(col, line, data[key], sheet)
                    else:
                        for k in data[key][0]:
                            write_arr_cell(col, line, [kl[k] for kl in data[key]], sheet)
                            col += 1
                except Exception as er:
                    print(er)
                    col += 1
            else:
                write_cell(set_cell(col, line), data[key], sheet)
                col += 1
        col = 1
        line += m_var

    xfile.save(file_name)


def write_one_unit(sheet, col, num, unit, region, title):
    if 'metro' not in unit:
        return 0
    count = max([len(unit['price_list']), len(unit['stuff_list']), len(unit['promos']), len(unit['tags'])])
    sim_keys = ['rating', 'phone', 'address', 'work_time', 'price', 'witget', 'site', 'its route']
    dict_keys = ['price_list', 'promos', 'stuff_list']
    arr_keys = ['tags', 'metro', 'rayons', 'social_links']
    c = 2
    n = num
    write_arr_cell(0, n, [reg_keys[region] for j in range(count)], sheet)
    write_arr_cell(1, n, [title.replace('_', ' ').replace('&nbsp;', ' ').replace('&amp;', ' ') for j in range(count)], sheet)
    for key in unit:
        if key in sim_keys:
            write_arr_cell(c, n, [unit[key] for j in range(count)], sheet)
            c += 1
        if key in dict_keys:
            res = []
            for i in unit[key]:
                st = ''
                for s_k in i:
                    st = st + ' ' + i[s_k]
                res.append(st)
            write_arr_cell(c, n, res, sheet)
            c += 1
        if key in arr_keys:
            write_arr_cell(c, n, unit[key], sheet)
            c += 1

    return count


if __name__ == "__main__":

    file_name = 'exele_files/{0}.xlsx'.format('bup_list')
    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet(name='Page1')
    worksheet.set_column('B:B', 45)
    worksheet.set_column('D:D', 30)
    worksheet.set_column('E:H', 50)
    worksheet.set_column('I:I', 100)
    worksheet.set_column('J:K', 100)
    workbook.close()
    headers = ['Город', 'Название', 'Рейтинг', 'Телефон', 'Адрес', 'Метро', 'Район', 'Время работы', 'Ценовая категория', 'Првйс лист', 'Сотрудники', 'Сайт', 'Сеть салонов', 'Соц сети', 'Акции', 'Теги', 'Виджет']
    xfile = openpyxl.load_workbook(file_name)
    sheet = xfile.get_sheet_by_name('Page1')

    line = 1
    col = 0
    for h in headers:
        write_cell(set_cell(col,1), h, sheet)
        col += 1
    col = 0
    line = 2
    # regions = ['msk', 'nsk', 'spb', 'ekb', 'vn', 'barnaul', 'arkhangelsk', 'astrakhan', 'chelyabinsk', 'kaliningrad', 'vladivostok', 'biysk', 'blag', 'bratsk', 'bryansk', 'vladimir', 'volgograd', 'vologda', 'voronezh', 'gorno-altaisk', 'ivanovo', 'izhevsk', 'irkutsk', 'kazan', 'kaluga', 'kirov', 'komsomolsk', 'kostroma', 'krasnodar', 'kurgan', 'kursk', 'voronezh', 'lipetsk', 'magnitogorsk', 'makhachkala', 'murmansk', 'chelny', 'nizhnevartovsk', 'nn', 'tagil', 'novokuznetsk', 'novorossiysk', 'omsk', 'orenburg', 'orel', 'penza', 'perm', 'pskov', 'rostov', 'samara', 'saransk', 'saratov', 'smolensk', 'sochi', 'surgut', 'syktyvkar', 'tambov', 'tver', 'togliatti', 'tomsk', 'tula', 'tyumen', 'ulan-ude', 'ulyanovsk', 'ufa', 'khabarovsk', 'cheboksary', 'chita', 'yuzhnosakhalinsk', 'yakutsk', 'yaroslavl']
    regions = ['msk']
    for r in regions:
        direction = 'beauty_pack/{0}'.format(r)
        i = 1
        for (dirpath, dirnames, filenames) in walk(direction):
            for but in filenames:
                if '.json' in but:
                    i += 1
                    try:
                        f = open('beauty_pack/{0}/{1}'.format(r, but), 'r').read()
                        line += write_one_unit(sheet, col, line, json.loads(f), r, but.replace('.json', ''))
                        print(but, 'ok')
                    except Exception as er:
                        print(but, er, 'bad')
                        line += 0
        print("all ", i)
    xfile.save(file_name)

